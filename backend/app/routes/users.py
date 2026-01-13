from flask import Blueprint, request, jsonify
from app.database import db
from app.models import Activity, Run, User, Challenge, ChallengeParticipant
from flask_jwt_extended import jwt_required, get_jwt_identity
from datetime import datetime
import re
import io
from openpyxl import load_workbook

users_bp = Blueprint('users', __name__)

@users_bp.route('/badges', methods=['GET'])
@jwt_required()
def get_user_badges():
    """Get badge counts (gold, silver, bronze) for the current user"""
    try:
        user_id_str = get_jwt_identity()
        user_id = int(user_id_str) if isinstance(user_id_str, str) else user_id_str
    except Exception as e:
        response = jsonify({'error': 'Invalid or expired token', 'details': str(e)})
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response, 401
    
    # Get all challenges the user has participated in
    participants = ChallengeParticipant.query.filter_by(user_id=user_id).all()
    
    gold_count = 0
    silver_count = 0
    bronze_count = 0
    
    # For each challenge, calculate the user's rank
    for participant in participants:
        challenge = Challenge.query.get(participant.challenge_id)
        if not challenge:
            continue
        
        # Skip challenges that haven't ended yet (only count final results)
        now = datetime.utcnow()
        if challenge.end_date > now:
            continue
        
        # Get all participants for this challenge
        all_participants = ChallengeParticipant.query.filter_by(challenge_id=challenge.id).all()
        
        # Sort by progress (descending for distance/time, ascending for fastest_5k)
        if challenge.challenge_type == 'fastest_5k':
            # For fastest 5K, lowest time wins
            all_participants.sort(key=lambda p: p.progress_value if p.progress_value > 0 else float('inf'))
        else:
            # For distance/time challenges, highest progress wins
            all_participants.sort(key=lambda p: p.progress_value, reverse=True)
        
        # Find user's rank
        user_rank = None
        for rank, p in enumerate(all_participants, 1):
            if p.user_id == user_id:
                user_rank = rank
                break
        
        # Count badges (only for top 3)
        if user_rank == 1:
            gold_count += 1
        elif user_rank == 2:
            silver_count += 1
        elif user_rank == 3:
            bronze_count += 1
    
    response = jsonify({
        'gold': gold_count,
        'silver': silver_count,
        'bronze': bronze_count
    })
    response.headers.add('Access-Control-Allow-Origin', '*')
    return response, 200

@users_bp.route('/activity-feed/<int:club_id>', methods=['GET'])
@jwt_required()
def get_activity_feed(club_id):
    try:
        user_id_str = get_jwt_identity()
        user_id = int(user_id_str) if isinstance(user_id_str, str) else user_id_str
    except Exception as e:
        response = jsonify({'error': 'Invalid or expired token', 'details': str(e)})
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response, 401
    
    # Only return run activities (tracked manually or via GPS)
    activities = Activity.query.filter_by(
        club_id=club_id,
        activity_type='run'
    ).order_by(Activity.created_at.desc()).limit(50).all()
    
    activities_data = []
    for activity in activities:
        # Try to find the related run by parsing description
        run_id = None
        run_data = None
        if activity.activity_type == 'run':
            # Parse description like "John ran 5.00 km at 10.00 km/h"
            match = re.search(r'ran ([\d.]+) km at ([\d.]+) km/h', activity.description)
            if match:
                distance = float(match.group(1))
                speed = float(match.group(2))
                # Find the run that matches this distance and speed (within tolerance)
                # Also match by date - find runs created around the same time as the activity
                runs = Run.query.filter_by(user_id=activity.user_id).order_by(Run.date.desc()).all()
                
                # First try exact match with time proximity
                for run in runs:
                    if abs(run.distance_km - distance) < 0.01 and abs(run.speed_kmh - speed) < 0.01:
                        # Check if the run date is close to activity creation time (within 24 hours)
                        time_diff = abs((run.date - activity.created_at).total_seconds())
                        if time_diff < 86400:  # Within 24 hours
                            run_id = run.id
                            run_data = {
                                'distance_km': run.distance_km,
                                'duration_minutes': run.duration_minutes,
                                'speed_kmh': run.speed_kmh,
                                'notes': run.notes,
                                'date': run.date.isoformat()
                            }
                            break
                
                # If no exact match found, try matching by distance and speed with relaxed tolerance
                if not run_data:
                    for run in runs:
                        if abs(run.distance_km - distance) < 0.1 and abs(run.speed_kmh - speed) < 0.5:
                            # Check time proximity (within 24 hours)
                            time_diff = abs((run.date - activity.created_at).total_seconds())
                            if time_diff < 86400:  # Within 24 hours
                                run_id = run.id
                                run_data = {
                                    'distance_km': run.distance_km,
                                    'duration_minutes': run.duration_minutes,
                                    'speed_kmh': run.speed_kmh,
                                    'notes': run.notes,
                                    'date': run.date.isoformat()
                                }
                                break
                
                # Last resort: match by distance only if created within 24 hours
                if not run_data:
                    for run in runs:
                        if abs(run.distance_km - distance) < 0.5:
                            time_diff = abs((run.date - activity.created_at).total_seconds())
                            if time_diff < 86400:  # Within 24 hours
                                run_id = run.id
                                run_data = {
                                    'distance_km': run.distance_km,
                                    'duration_minutes': run.duration_minutes,
                                    'speed_kmh': run.speed_kmh,
                                    'notes': run.notes,
                                    'date': run.date.isoformat()
                                }
                                break
        
        activities_data.append({
            'id': activity.id,
            'type': activity.activity_type,
            'description': activity.description,
            'user_name': activity.user.name,
            'user_id': activity.user_id,
            'created_at': activity.created_at.isoformat(),
            'run_id': run_id,
            'run_data': run_data
        })
    
    response = jsonify(activities_data)
    response.headers.add('Access-Control-Allow-Origin', '*')
    return response, 200

@users_bp.route('/activities/<int:activity_id>', methods=['PUT'])
@jwt_required()
def update_activity(activity_id):
    try:
        user_id_str = get_jwt_identity()
        user_id = int(user_id_str) if isinstance(user_id_str, str) else user_id_str
    except Exception as e:
        response = jsonify({'error': 'Invalid or expired token', 'details': str(e)})
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response, 401
    
    activity = Activity.query.get(activity_id)
    if not activity:
        response = jsonify({'error': 'Activity not found'})
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response, 404
    
    # Check if user owns this activity
    if activity.user_id != user_id:
        response = jsonify({'error': 'You can only edit your own activities'})
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response, 403
    
    data = request.get_json()
    
    if not data:
        response = jsonify({'error': 'No data provided'})
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response, 400
    
    try:
        # If it's a run activity, update the underlying run
        if activity.activity_type == 'run' and data.get('run_data'):
            run_data = data['run_data']
            
            # Find the related run
            match = re.search(r'ran ([\d.]+) km at ([\d.]+) km/h', activity.description)
            if match:
                distance = float(match.group(1))
                speed = float(match.group(2))
                runs = Run.query.filter_by(user_id=user_id).all()
                run = None
                for r in runs:
                    if abs(r.distance_km - distance) < 0.01 and abs(r.speed_kmh - speed) < 0.01:
                        run = r
                        break
                
                if run:
                    # Update run
                    if 'distance_km' in run_data:
                        run.distance_km = float(run_data['distance_km'])
                    if 'duration_minutes' in run_data:
                        run.duration_minutes = float(run_data['duration_minutes'])
                    if 'notes' in run_data:
                        run.notes = run_data['notes']
                    if 'date' in run_data:
                        try:
                            from datetime import datetime
                            run.date = datetime.fromisoformat(run_data['date'].replace('Z', '+00:00'))
                        except:
                            pass
                    
                    # Recalculate speed
                    if 'distance_km' in run_data or 'duration_minutes' in run_data:
                        run.speed_kmh = (run.distance_km / run.duration_minutes) * 60 if run.duration_minutes > 0 else 0
                    
                    # Update activity description
                    user = User.query.get(user_id)
                    activity.description = f'{user.name} ran {run.distance_km:.2f} km at {run.speed_kmh:.2f} km/h'
                    
                    db.session.commit()
                    
                    response = jsonify({
                        'id': activity.id,
                        'description': activity.description,
                        'created_at': activity.created_at.isoformat(),
                        'run_data': {
                            'distance_km': run.distance_km,
                            'duration_minutes': run.duration_minutes,
                            'speed_kmh': run.speed_kmh,
                            'notes': run.notes,
                            'date': run.date.isoformat()
                        }
                    })
                    response.headers.add('Access-Control-Allow-Origin', '*')
                    return response, 200
        
        # If no run update, just update description
        if 'description' in data:
            activity.description = data['description']
            db.session.commit()
        
        response = jsonify({
            'id': activity.id,
            'description': activity.description,
            'created_at': activity.created_at.isoformat()
        })
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response, 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"Error updating activity: {e}")
        traceback.print_exc()
        response = jsonify({'error': f'Failed to update activity: {str(e)}'})
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type, Authorization')
        return response, 500

@users_bp.route('/activities/<int:activity_id>', methods=['DELETE'])
@jwt_required()
def delete_activity(activity_id):
    try:
        user_id_str = get_jwt_identity()
        user_id = int(user_id_str) if isinstance(user_id_str, str) else user_id_str
    except Exception as e:
        response = jsonify({'error': 'Invalid or expired token', 'details': str(e)})
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response, 401
    
    activity = Activity.query.get(activity_id)
    if not activity:
        response = jsonify({'error': 'Activity not found'})
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response, 404
    
    # Check if user owns this activity
    if activity.user_id != user_id:
        response = jsonify({'error': 'You can only delete your own activities'})
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response, 403
    
    try:
        # Delete the activity (keep the run if it exists)
        db.session.delete(activity)
        db.session.commit()
        
        response = jsonify({'message': 'Activity deleted successfully'})
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response, 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"Error deleting activity: {e}")
        traceback.print_exc()
        response = jsonify({'error': f'Failed to delete activity: {str(e)}'})
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type, Authorization')
        return response, 500

@users_bp.route('/bulk-import', methods=['POST'])
@jwt_required()
def bulk_import_users():
    """Import users from an Excel file"""
    try:
        user_id_str = get_jwt_identity()
        user_id = int(user_id_str) if isinstance(user_id_str, str) else user_id_str
    except Exception as e:
        response = jsonify({'error': 'Invalid or expired token', 'details': str(e)})
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response, 401
    
    # Check if file is present
    if 'file' not in request.files:
        response = jsonify({'error': 'No file provided'})
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response, 400
    
    file = request.files['file']
    
    if file.filename == '':
        response = jsonify({'error': 'No file selected'})
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response, 400
    
    # Check file extension
    if not file.filename.endswith(('.xlsx', '.xls')):
        response = jsonify({'error': 'Invalid file type. Please upload an Excel file (.xlsx or .xls)'})
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response, 400
    
    try:
        # Read Excel file using openpyxl
        file_content = file.read()
        workbook = load_workbook(io.BytesIO(file_content), data_only=True)
        sheet = workbook.active
        
        # Get header row (first row)
        if sheet.max_row < 2:
            response = jsonify({'error': 'Excel file is empty or has no data rows'})
            response.headers.add('Access-Control-Allow-Origin', '*')
            return response, 400
        
        # Read header row and normalize column names (case-insensitive)
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        column_indices = {}
        required_columns = ['name', 'email', 'password']
        
        # Map normalized column name to index (0-indexed)
        # Note: address column is skipped to avoid database errors
        for idx, cell_value in enumerate(header_row):
            if cell_value:
                col_name = str(cell_value).strip().lower()
                if col_name in ['name', 'email', 'password']:
                    column_indices[col_name] = idx
        
        # Validate required columns
        missing_columns = [col for col in required_columns if col not in column_indices]
        
        if missing_columns:
            response = jsonify({
                'error': f'Missing required columns: {", ".join(missing_columns)}',
                'required_columns': required_columns,
                'found_columns': [col for col in column_indices.keys()]
            })
            response.headers.add('Access-Control-Allow-Origin', '*')
            return response, 400
        
        # Process each row
        created_users = []
        skipped_users = []
        errors = []
        
        # Start from row 2 (skip header)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Get values (handle None/empty values)
                # Note: address field is skipped to avoid database column errors
                name = str(row[column_indices['name']]).strip() if column_indices['name'] < len(row) and row[column_indices['name']] is not None else None
                email = str(row[column_indices['email']]).strip() if column_indices['email'] < len(row) and row[column_indices['email']] is not None else None
                password = str(row[column_indices['password']]).strip() if column_indices['password'] < len(row) and row[column_indices['password']] is not None else None
                
                # Validate required fields
                if not name or not email or not password:
                    errors.append({
                        'row': row_idx,
                        'error': 'Missing required field (name, email, or password)'
                    })
                    continue
                
                # Validate email format
                if '@' not in email:
                    errors.append({
                        'row': row_idx,
                        'error': f'Invalid email format: {email}'
                    })
                    continue
                
                # Check if user already exists
                existing_user = User.query.filter_by(email=email).first()
                if existing_user:
                    skipped_users.append({
                        'row': row_idx,
                        'email': email,
                        'reason': 'User already exists'
                    })
                    continue
                
                # Create new user (address field skipped to avoid database errors)
                user = User(
                    email=email,
                    name=name
                )
                user.set_password(password)
                
                db.session.add(user)
                created_users.append({
                    'name': name,
                    'email': email
                })
                
            except Exception as e:
                errors.append({
                    'row': row_idx,
                    'error': f'Error processing row: {str(e)}'
                })
                continue
        
        # Commit all users
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            response = jsonify({
                'error': f'Failed to create users: {str(e)}',
                'created_count': len(created_users),
                'skipped_count': len(skipped_users),
                'error_count': len(errors)
            })
            response.headers.add('Access-Control-Allow-Origin', '*')
            return response, 500
        
        response = jsonify({
            'message': 'Users imported successfully',
            'created_count': len(created_users),
            'skipped_count': len(skipped_users),
            'error_count': len(errors),
            'created_users': created_users,
            'skipped_users': skipped_users,
            'errors': errors
        })
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response, 200
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        response = jsonify({'error': f'Failed to process Excel file: {str(e)}'})
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response, 500
